# -*- coding:utf-8 -*-
"""
@Author: Komorebi 
"""
import excel.utils as utils
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)


SUBJECT = [
    'subject player',
    'subject hero',
]

OBJECT = [
    'object player',
    'object hero',
]

SUPPLEMENT = [
    'ability',
    'critical kill',
]

ASSIST = [
    'a player 1',
    'a hero 1',
    'a player 2',
    'a hero 2',
    'a player 3',
    'a hero 3',
    'a player 4',
    'a hero 4',
    'a player 5',
    'a hero 5',
]

TITLE = [
    'time',
    'action',
] + SUBJECT + OBJECT + SUPPLEMENT + ASSIST

PRINT_DATA_FORMAT = {t: i+1 for i, t in enumerate(TITLE + ['comments'])}

TITLE_TOP = [
    '',
    '',
    'subject',
    '',
    'object',
    '',
    '',
    '',
    'assist 1',
    '',
    'assist 2',
    '',
    'assist 3',
    '',
    'assist 4',
    '',
    'assist 5',
    '',
    'comments',
]

DIMENSIONS = {
    'time': 'A',
    'action': 'B',
    'subject player': 'C',
    'subject hero': 'D',
    'object player': 'E',
    'object hero': 'F',
    'ability': 'G',
    'critical kill': 'H',
    'a player 1': 'I',
    'a hero 1': 'J',
    'a player 2': 'K',
    'a hero 2': 'L',
    'a player 3': 'M',
    'a hero 3': 'N',
    'a player 4': 'O',
    'a hero 4': 'P',
    'a player 5': 'Q',
    'a hero 5': 'R',
    'comments': 'S',
}

TITLE_TOP_MERGE_CELL = {
    '': None,
    'subject': '{}1:{}1'.format(DIMENSIONS['subject player'], DIMENSIONS['subject hero']),
    'object': '{}1:{}1'.format(DIMENSIONS['object player'], DIMENSIONS['object hero']),
    'assist 1': '{}1:{}1'.format(DIMENSIONS['a player 1'], DIMENSIONS['a hero 1']),
    'assist 2': '{}1:{}1'.format(DIMENSIONS['a player 2'], DIMENSIONS['a hero 2']),
    'assist 3': '{}1:{}1'.format(DIMENSIONS['a player 3'], DIMENSIONS['a hero 3']),
    'assist 4': '{}1:{}1'.format(DIMENSIONS['a player 4'], DIMENSIONS['a hero 4']),
    'assist 5': '{}1:{}1'.format(DIMENSIONS['a player 5'], DIMENSIONS['a hero 5']),
    'comments': '{}1:{}2'.format(DIMENSIONS['comments'], DIMENSIONS['comments']),
}

PLAYER_WIDTH_CONFIG = {DIMENSIONS['a player {}'.format(i)]: 18 for i in range(1, 6)}
HERO_WIDTH_CONFIG = {DIMENSIONS['a hero {}'.format(i)]: 16 for i in range(1, 6)}
CELL_WIDTH_CONFIG = {
    DIMENSIONS['time']: 16.5,
    DIMENSIONS['action']: 20,
    DIMENSIONS['subject player']: 18,
    DIMENSIONS['subject hero']: 16,
    DIMENSIONS['object player']: 18,
    DIMENSIONS['object hero']: 16,
    DIMENSIONS['ability']: 14.5,
    DIMENSIONS['critical kill']: 14,
    DIMENSIONS['comments']: 45.5,
}
CELL_WIDTH_CONFIG.update(PLAYER_WIDTH_CONFIG)
CELL_WIDTH_CONFIG.update(HERO_WIDTH_CONFIG)


ABILITY_FORMAT = {
    0: 'Plain Attack',
    1: 'Shift',
    2: 'E',
    3: 'Ultmate 1',
    4: 'Ultmate 2',
    5: 'RMB',
    6: 'Passive',
}


def _cell_style():
    d = {
        'font': {
            'name': 'Microsoft YaHei',
            'size': 12,
            'bold': True,
            'vertAlign': 'baseline',
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center',
            'wrap_text': True,
        },
        'fill': {},
    }
    return d


def set_action(obj):
    player1, player2 = obj.player1, obj.player2
    if player1['chara'] == 'mercy' and player1['team'] == player2['team']:
        return 'Resurrect'
    elif player2['chara'] == 'meka':
        return 'Demech'
    else:
        return 'Eliminate'


def set_comments(action):
    table = {
        'Resurrect': 'Resurrect',
        'Demech': 'MEKA destroyed',
        'Eliminate': '',
    }
    return table[action]


class Config(object):
    cell_width = CELL_WIDTH_CONFIG
    cell_height = 18
    cell_style = _cell_style()
    merge_cell = TITLE_TOP_MERGE_CELL
    format = PRINT_DATA_FORMAT
    team_colors = {}
    title_top = TITLE_TOP
    title = TITLE
    ability = ABILITY_FORMAT


class Save:
    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def _set_cell_style(cell, title):
        style = Config.cell_style
        cell.font = Font(**style['font'])
        cell.alignment = Alignment(**style['alignment'])
        if title in style['fill'].keys():
            cell.fill = PatternFill(**style['fill'][title])

    def _set_cells_style(self):
        max_row, max_column = self.sheet.max_row + 1, self.sheet.max_column + 1
        for r in range(1, max_row):
            self.sheet.row_dimensions[r].height = Config.cell_height
            for c in range(1, max_column):
                letter = get_column_letter(c)
                title = '{}{}'.format(letter, str(r))
                self._set_cell_style(self.sheet[title], title)
                self.sheet.column_dimensions[letter].width = Config.cell_width[letter]

    def _set_cells_merge(self):
        merge_config = Config.merge_cell
        for key in Config.title_top:
            if merge_config[key] is not None:
                self.sheet.merge_cells(merge_config[key])

    def save(self):
        self._set_cells_style()
        self._set_cells_merge()


class Sheet:
    def __init__(self, wb, game):
        self.game = game
        self.sheet = wb['sheet1']
        self.sheet.append(Config.title_top)
        self.sheet.append(Config.title)
        Config.team_colors[self.game.team_names['left']] = utils.to_hex(self.game.team_colors['left'])
        Config.team_colors[self.game.team_names['right']] = utils.to_hex(self.game.team_colors['right'])

        # 上一帧的大招状况
        self.ultimate_status = {i: False for i in range(1, 13)}

        self.previous_chara = [player.chara for player in game.frames[0].players]
        self.next_chara = [player.chara for player in game.frames[1].players]

    def _new_data(self, data):
        """
        从数据中提取出想要的信息，并加入到 Config 中
        """
        if '_$color' in data.keys():
            color = data.pop('_$color')
            for k, v in color.items():
                cell = '{}{}'.format(DIMENSIONS[k], self.sheet.max_row + 1)
                Config.cell_style['fill'][cell] = {
                    'fill_type': 'solid',
                    'fgColor': v,
                }
        return data

    @staticmethod
    def _format(data):
        """
        格式化输出信息
        :param data: {"字段名": value, ...}
        :return: [value, ...]
        """
        data['time'] = utils.time_format(data['time'])
        format_spec = Config.format
        result = [''] * (len(Config.title) + 1)
        for k, v in format_spec.items():
            if k in ['object hero', 'subject hero'] and k in data:
                data[k] = utils.capitalize(data[k])
            if k in data:
                result[v-1] = data[k]

        for i, s in enumerate(result):
            if s == 'empty':
                result[i] = ''
        return result

    def _append(self, **kwargs):
        """
        将数据添加到 sheet 中
        """
        kwargs = self._new_data(kwargs)
        self.sheet.append(self._format(kwargs))

    def new(self):
        frames = self.game.frames
        for i, frame in enumerate(frames):
            self._killfeed_append(frame.killfeeds, frame.time)
            self._ultimate_append(frame.players, frame.time)
            self._switch_hero_append(frame.players, frame.time, i)
        self.save()

    def _killfeed_append(self, killfeeds, time):
        """
        将 killfeed 中的属性替换成 change 中的，并添加到表中
        """
        for obj in killfeeds:
            d = {}
            player1, player2 = obj.player1, obj.player2
            d['time'] = time
            # 这里先写死
            d['action'] = set_action(obj)
            d['comments'] = set_comments(d['action'])
            d['ability'] = Config.ability[obj.ability]
            d['subject hero'] = player1['chara']
            d['subject player'] = player1['player']
            d['object hero'] = player2['chara']
            d['object player'] = player2['player']
            d['_$color'] = {}
            for i, p in enumerate([player1, player2]):
                if p['player'] != 'empty':
                    if i == 0:
                        d['_$color']['subject player'] = Config.team_colors[player1['team']]
                    else:
                        d['_$color']['object player'] = Config.team_colors[player2['team']]

            for i, assist in enumerate(obj.assists):
                d['a player {}'.format(i + 1)] = assist['player']
                d['a hero {}'.format(i + 1)] = utils.capitalize(assist['chara'])
                if assist['player'] != 'empty':
                    d['_$color']['a player {}'.format(i + 1)] = Config.team_colors[assist['team']]
            self._append(**d)

    def _ultimate_append(self, players, time):
        """
        判断此时的大招是否为刚充满/释放，并将刚充满/释放大招的选手名输出
        """
        status = self.ultimate_status
        for player in players:
            d = {
                'time': time,
                'subject player': player.name,
                'subject hero': player.chara,
                '_$color': {
                    'subject player': Config.team_colors[player.team]
                }
            }
            index = player.index + 1
            if player.is_ult_ready and not status[index]:
                status[index] = True
                d['action'] = 'ult ready'
                self._append(**d)
            elif not player.is_ult_ready and status[index]:
                status[index] = False
                d['action'] = 'ult used'
                self._append(**d)
            else:
                break

    def _switch_hero_append(self, players, time, index):
        frames = self.game.frames
        length = len(frames)
        if index == 0:
            return
        elif index >= length - 2:
            return
        else:
            for i, player in enumerate(players):
                top_chara = self.previous_chara[i]
                next_chara = self.next_chara[i]
                if player.is_dead:
                    continue
                elif top_chara != player.chara:
                    if top_chara == next_chara:
                        self.next_chara[i] = frames[index + 2].players[i].chara
                        continue
                    elif player.chara == next_chara:
                        d = {
                            'time': time,
                            'action': 'hero switch',
                            'subject player': player.name,
                            'subject hero': player.chara,
                            'comments': 'Switch from {} to {}'.format(utils.capitalize(top_chara), utils.capitalize(player.chara)),
                            '_$color': {
                                'subject player': Config.team_colors[player.team],
                            }
                        }
                        self._append(**d)
                        self.previous_chara[i] = player.chara
                        self.next_chara[i] = frames[index + 1].players[i].chara

    def save(self):
        """
        对 sheet 中单元格应用样式并保存
        :return: None
        """
        Save(self.sheet).save()
