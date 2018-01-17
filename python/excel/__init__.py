"""
@Author: Komorebi 
"""
from openpyxl import Workbook
from excel.sheet1 import Sheet as Sheet1
from excel.sheet2 import Sheet as Sheet2


def create_sheet():
    wb = Workbook()
    wb.active.title = 'sheet1'
    wb.create_sheet('sheet2')
    return wb


class Excel(object):
    def __init__(self, game):
        self._wb = create_sheet()
        self.game = game
        self.sheet1 = Sheet1(self._wb, game)
        self.sheet2 = Sheet2(self._wb, game)

    def save(self, filename='text', postfix='xlsx'):
        self.sheet1.new()
        self.sheet2.new()
        self._wb.save('{}\{}.{}'.format(self.game.output_path, filename, postfix))
