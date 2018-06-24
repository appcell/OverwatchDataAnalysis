"""
@Author: Komorebi 
"""
from openpyxl import Workbook
from .sheet1 import Sheet as Sheet1
from .sheet2 import Sheet as Sheet2
from .sheet3 import Sheet as Sheet3


def create_sheet():
    wb = Workbook()
    wb.active.title = 'sheet1'
    wb.create_sheet('sheet2')
    wb.create_sheet('sheet3')
    return wb


class Excel(object):
    def __init__(self, game):
        self._wb = create_sheet()
        self.game = game
        self.sheet1 = Sheet1(self._wb, game)
        self.sheet2 = Sheet2(self._wb, game)
        self.sheet3 = Sheet3(self._wb, game)

    def save(self):
        self.sheet1.new()
        self.sheet2.new()
        self.sheet3.new()
        self._wb.save(self.game.output_path)
        return self.json()

    def json(self):
        sheet1_data = self.sheet1.json()
        sheet2_data = self.sheet2.json()
        sheet3_data = self.sheet3.json()
        return [sheet1_data, sheet2_data, sheet3_data]
