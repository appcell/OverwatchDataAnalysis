"""
@Author: Komorebi 
"""
from json import dump
from . import utils
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)


SUBJECT = [
    'subject player',
    'subject chara',
]

OBJECT = [
    'object player',
    'object chara',
]

SUPPLEMENT = [
    'ability',
    'critical kill',
]

ASSIST = [
    'a player 1',
    'a hero 1',
    'a player 2',
    'a hero 2',
    'a player 3',
    'a hero 3',
    'a player 4',
    'a hero 4',
    'a player 5',
    'a hero 5',
]

TITLE = [
    'time',
    'action',
] + SUBJECT + OBJECT + SUPPLEMENT + ASSIST

PRINT_DATA_FORMAT = {t: i+1 for i, t in enumerate(TITLE + ['comments'])}

TITLE_TOP = [
    '',
    '',
    'subject',
    '',
    'object',
    '',
    '',
    '',
    'assist 1',
    '',
    'assist 2',
    '',
    'assist 3',
    '',
    'assist 4',
    '',
    'assist 5',
    '',
    'comments',
]

DIMENSIONS = {
    'time': 'A',
    'action': 'B',
    'subject player': 'C',
    'subject chara': 'D',
    'object player': 'E',
    'object chara': 'F',
    'ability': 'G',
    'critical kill': 'H',
    'a player 1': 'I',
    'a hero 1': 'J',
    'a player 2': 'K',
    'a hero 2': 'L',
    'a player 3': 'M',
    'a hero 3': 'N',
    'a player 4': 'O',
    'a hero 4': 'P',
    'a player 5': 'Q',
    'a hero 5': 'R',
    'comments': 'S',
}

TITLE_TOP_MERGE_CELL = {
    '': None,
    'subject': '{}1:{}1'.format(DIMENSIONS['subject player'], DIMENSIONS['subject chara']),
    'object': '{}1:{}1'.format(DIMENSIONS['object player'], DIMENSIONS['object chara']),
    'assist 1': '{}1:{}1'.format(DIMENSIONS['a player 1'], DIMENSIONS['a hero 1']),
    'assist 2': '{}1:{}1'.format(DIMENSIONS['a player 2'], DIMENSIONS['a hero 2']),
    'assist 3': '{}1:{}1'.format(DIMENSIONS['a player 3'], DIMENSIONS['a hero 3']),
    'assist 4': '{}1:{}1'.format(DIMENSIONS['a player 4'], DIMENSIONS['a hero 4']),
    'assist 5': '{}1:{}1'.format(DIMENSIONS['a player 5'], DIMENSIONS['a hero 5']),
    'comments': '{}1:{}2'.format(DIMENSIONS['comments'], DIMENSIONS['comments']),
}

PLAYER_WIDTH_CONFIG = {DIMENSIONS['a player {}'.format(i)]: 18 for i in range(1, 6)}
HERO_WIDTH_CONFIG = {DIMENSIONS['a hero {}'.format(i)]: 16 for i in range(1, 6)}
CELL_WIDTH_CONFIG = {
    DIMENSIONS['time']: 16.5,
    DIMENSIONS['action']: 20,
    DIMENSIONS['subject player']: 18,
    DIMENSIONS['subject chara']: 16,
    DIMENSIONS['object player']: 18,
    DIMENSIONS['object chara']: 16,
    DIMENSIONS['ability']: 14.5,
    DIMENSIONS['critical kill']: 14,
    DIMENSIONS['comments']: 45.5,
}
CELL_WIDTH_CONFIG.update(PLAYER_WIDTH_CONFIG)
CELL_WIDTH_CONFIG.update(HERO_WIDTH_CONFIG)


ABILITY_FORMAT = {
    0: 'Plain Attack',
    1: 'Shift',
    2: 'E',
    3: 'Ultimate 1',
    4: 'Ultimate 2',
    5: 'Right Click',
    6: 'Passive',
}


def _cell_style():
    """
    sheet1 中 cell 的样式
    :return: None
    """
    d = {
        'font1': {
            'name': 'Microsoft YaHei',
            'size': 12,
            'bold': True,
            'vertAlign': 'baseline',
        },
        'font2': {
            'name': 'Microsoft YaHei',
            'size': 12,
            'bold': True,
            'vertAlign': 'baseline',
            'color': 'FFFFFF',
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center',
            'wrap_text': True,
        },
        'fill': {},
    }
    return d


def set_action(obj):
    """
    通过对 player1、player2 相关的判断来获取 action
    :param obj: killfeed类
    :return: action
    """
    player1, player2 = obj.player1, obj.player2
    if player1['chara'] == 'mercy' and player1['team'] == player2['team']:
        return 'Resurrect'
    elif player2['chara'] == 'meka':
        return 'Demech'
    elif player1['chara'] == 'empty' and player1['team'] == 'empty' and player2['chara'] is not None:
        return "Suicide"
    else:
        return 'Eliminate'


def set_comments(action):
    """
    通过 action 来将相应的信息转换成 comment
    :return: comment
    """
    table = {
        'Resurrect': 'Resurrect',
        'Demech': 'MEKA destroyed',
        'Eliminate': '',
        'Suicide': ''
    }
    return table[action]


def get_player_name(player_index, player_names):
    """
    通过编号获取玩家姓名
    :param player_index: 玩家的编号
    :param player_names: 存储了12个玩家姓名的 list
    :return: player_name
    """
    return player_names[player_index]


def get_player_team_index(player_index):
    """
    获取玩家所在的队伍编号
    :param player_index: 玩家所处的队伍名
    :return: 0 or 1
    """
    return 0 if player_index < 6 else 1


class Config(object):
    cell_width = CELL_WIDTH_CONFIG
    cell_height = 18
    cell_style = _cell_style()
    merge_cell = TITLE_TOP_MERGE_CELL
    format = PRINT_DATA_FORMAT
    team_colors = {}
    title_top = TITLE_TOP
    title = TITLE
    ability = ABILITY_FORMAT
    peculiar_cell = []


class Save:
    def __init__(self, sheet, data):
        self.sheet = sheet
        self.data = data

    @staticmethod
    def _set_cell_style(cell, title):
        """
        将样式应用到 cell 中
        :param cell: sheet 中的 cell
        :param title: cell 坐标， 如 A1、B2 ..
        :return: None
        """
        style = Config.cell_style
        cell.font = Font(**style['font2']) if title in Config.peculiar_cell else Font(**style['font1'])
        cell.alignment = Alignment(**style['alignment'])
        if title in style['fill'].keys():
            cell.fill = PatternFill(**style['fill'][title])

    def _set_cells_style(self):
        """
        将样式应用到所有 cell 中 
        """
        max_row, max_column = self.sheet.max_row + 1, self.sheet.max_column + 1
        for r in range(1, max_row):
            self.sheet.row_dimensions[r].height = Config.cell_height
            for c in range(1, max_column):
                letter = get_column_letter(c)
                title = letter + str(r)
                self._set_cell_style(self.sheet[title], title)
                self.sheet.column_dimensions[letter].width = Config.cell_width[letter]

    def _set_cells_merge(self):
        """
        合并 cell 
        """
        merge_config = Config.merge_cell
        for key in Config.title_top:
            if merge_config[key] is not None:
                self.sheet.merge_cells(merge_config[key])

    def _append(self):
        """
        将 self.data 中的数据按照时间排序, 并导入到sheet中 
        """
        self.data.sort(key=lambda x: x['_time'])
        for data in self.data:
            if '_$color' in data.keys():
                color = data.pop('_$color')
                for k, v in color.items():
                    title = DIMENSIONS[k] + str(self.sheet.max_row + 1)
                    c, b = v
                    if b:
                        Config.peculiar_cell.append(title)
                    Config.cell_style['fill'][title] = {
                        'fill_type': 'solid',
                        'fgColor': c,
                    }
            self.sheet.append(self._format(data))

    @staticmethod
    def _format(data):
        """
        格式化输出信息
        :param data: {"字段名": value, ...}
        :return: [value, ...]
        """
        data['time'] = utils.time_format(data['time'])
        format_spec = Config.format
        result = [''] * (len(Config.title) + 1)
        for k, v in format_spec.items():
            if k in ['object chara', 'subject chara'] and k in data:
                data[k] = utils.chara_capitalize(data[k])
            if k in data:
                result[v-1] = data[k]

        for i, s in enumerate(result):
            if s == 'empty' or s == 'Empty':
                result[i] = ''
        return result

    def save(self):
        self._append()
        self._set_cells_style()
        self._set_cells_merge()
        self.sheet.freeze_panes = self.sheet['B3']


class Sheet:
    def __init__(self, wb, game):
        self.game = game
        self.sheet = wb['sheet1']
        # 初始化
        self.sheet.append(Config.title_top)
        self.sheet.append(Config.title)
        if self.game.team_colors is None:
            Config.team_colors[0] = utils.to_hex([255, 255, 255])
            Config.team_colors[1] = utils.to_hex([70, 70, 70])
        Config.team_colors[0] = utils.to_hex(self.game.team_colors[0])
        Config.team_colors[1] = utils.to_hex(self.game.team_colors[1])

        self.data = []

    @staticmethod
    def _new_data(data):
        """
        为字典添加 _time key，用来排序
        """
        data['_time'] = data['time']
        return data

    def _append(self, **kwargs):
        """
        将所有数据添加到 self.data 中，以便后续处理
        """
        new_data = self._new_data(kwargs)
        self.data.append(new_data)

    def new(self):
        """
        遍历 frames， 提取相应的信息并保存
        """
        frames = self.game.frames
        for i, frame in enumerate(frames):
            self._switch_hero_append(frame.players, frame.time, i)
            self._killfeed_append(frame.killfeeds, frame.time)
            self._ultimate_append(frame.players, frame.time, i)
        self.save()

    def _killfeed_append(self, killfeeds, time):
        """
        往 sheet 加入击杀、自杀相关的信息。
        """
        for obj in killfeeds:
            d = {}
            player1, player2 = obj.player1, obj.player2
            d['time'] = time
            d['action'] = set_action(obj)
            d['comments'] = set_comments(d['action'])
            d['ability'] = Config.ability[obj.ability]
            d['subject chara'] = player1['chara']
            d['subject player'] = 'empty' if player1['player'] == -1 else self.game.name_players[player1['player']]
            d['object chara'] = player2['chara']
            d['object player'] = 'empty' if player2['player'] == -1 else self.game.name_players[player2['player']]
            if obj.is_headshot:
                d['critical kill'] = 'Y'
                d['PS'] = 'Head Shot'
            d['_$color'] = {}

            if d['subject player'] != 'empty':
                d['_$color']['subject player'] = Config.team_colors[player1['team']]

            if d['object player'] != 'empty':
                d['_$color']['object player'] = Config.team_colors[player2['team']]

            for i, assist in enumerate(obj.assists):
                d['a player {}'.format(i + 1)] = self.game.name_players[assist['player']]
                d['a hero {}'.format(i + 1)] = utils.chara_capitalize(assist['chara'])
                d['_$color']['a player {}'.format(i + 1)] = Config.team_colors[assist['team']]
            self._append(**d)

    def _ultimate_append(self, players, time, idx):
        """
        判断此时的大招是否为刚充满/释放，并将刚充满/释放大招的选手名输出
        """
        for index, player in enumerate(players):
            d = {
                'time': time,
                'subject player': get_player_name(player.index, self.game.name_players),
                'subject chara': player.chara,
                '_$color': {
                    'subject player': Config.team_colors[get_player_team_index(player.index)]
                }
            }
            previous_player = self.game.frames[idx - 1].players[index]
            if idx > 0:
                if not previous_player.is_ult_ready and player.is_ult_ready:
                    d['action'] = 'Ult ready'
                    self._append(**d)
                elif previous_player.is_ult_ready and not player.is_ult_ready:
                    d['action'] = 'Meka up' if player.dva_status == 1 and player.chara == 'dva' else 'Ult used'
                    self._append(**d)
            else:
                if player.is_ult_ready:
                    d['action'] = 'Ult ready'
                    self._append(**d)

    def _switch_hero_append(self, players, time, index):
        """
        通过与上一帧的英雄和下一帧的英雄的比对来判定玩家是否更换英雄

        :param players: 12个player类构成的 list ， 对应12个玩家
        :param time: 当前时间
        :param index: frames 中的第n个 frame
        """
        frames = self.game.frames
        if index == 0:
            return
        elif index >= len(frames) - 1:
            return
        previous_charas = frames[index - 1].players
        next_charas = frames[index + 1].players
        for i, player in enumerate(players):
            top_chara = previous_charas[i].chara
            next_chara = next_charas[i].chara
            if top_chara != player.chara:
                if top_chara == next_chara:
                    continue
                elif player.chara == next_chara:
                    d = {
                        'time': time - 1.5,
                        'action': 'Hero switch',
                        'subject player': get_player_name(player.index, self.game.name_players),
                        'subject chara': player.chara,
                        'comments': 'Switch from {} to {}'.format(utils.chara_capitalize(top_chara),
                                                                  utils.chara_capitalize(player.chara)),

                        '_$color': {
                            'subject player': Config.team_colors[get_player_team_index(player.index)],
                        }
                    }
                    self._append(**d)

    def save(self):
        """
        对 sheet 中单元格应用样式并保存
        :return: None
        """
        Save(self.sheet, self.data).save()

    def json(self):
        titles = {v: k for k, v in DIMENSIONS.items()}
        sheet_data = []
        for col, cells in enumerate(self.sheet.rows):
            if col < 2:
                continue
            data = {
                'subject': {
                    'team': '',
                },
                'object': {
                    'team': '',
                },
                'assist': {
                    '1': {},
                    '2': {},
                    '3': {},
                    '4': {},
                    '5': {},
                },
            }
            for cell in cells:
                title = titles[cell.column]
                if title in SUBJECT:
                    unit = title[8:]
                    data['subject'][unit] = cell.value
                    if unit == 'player' and cell.value != '':
                        data['subject']['team'] = self._get_team(cell)
                elif title in OBJECT:
                    unit = title[7:]
                    data['object'][unit] = cell.value
                    if unit == 'player' and cell.value != '':
                        data['object']['team'] = self._get_team(cell)
                elif title in ASSIST:
                    data['assist'][title[-1]][title[2:-2]] = cell.value
                else:
                    data[title] = cell.value
            sheet_data.append(data)
        return sheet_data

    def _get_team(self, cell):
        color = cell.fill.fgColor.__dict__['rgb'][2:]
        return self.game.team_names[0] if color == Config.team_colors[0][0] else self.game.team_names[1]
