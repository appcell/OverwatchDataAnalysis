# -*- coding:utf-8 -*-
"""
@Author: Rigel
"""
from json import dump
from . import utils
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)


TITLE = ['time'] + \
        ['player',
         'HP',
         'ult'] * 12

DIMENSIONS = {
    'time': ['A'],
    'players': ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W', 'Z', 'AC', 'AF', 'AI'],
    'hps': ['C', 'F', 'I', 'L', 'O', 'R', 'U', 'X', 'AA', 'AD', 'AG', 'AJ'],
    'ults': ['D', 'G', 'J', 'M', 'P', 'S', 'V', 'Y', 'AB', 'AE', 'AH', 'AK']
}

CELL_WIDTH_CONFIG = {
    'time': 16.5,
    'players': 18,
    'hps': 4.8,
    'ults': 4.8
}


def _cell_style():
    """
    sheet3 中 cell 的样式
    :return: None
    """
    d = {
        'font1': {
            'name': 'Microsoft YaHei',
            'size': 12,
            'bold': True,
            'vertAlign': 'baseline',
        },
        'font2': {
            'name': 'Microsoft YaHei',
            'size': 12,
            'bold': True,
            'vertAlign': 'baseline',
            'color': 'FFFFFF',
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center',
            'wrap_text': True,
        }
    }
    return d


class Config(object):
    cell_width = CELL_WIDTH_CONFIG
    dimensions = DIMENSIONS
    cell_height = 18
    cell_style = _cell_style()
    freeze_cell = 'B2'
    team_colors = {}
    title = TITLE
    peculiar_cell = []


class Save:
    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def _set_cell_style(cell, font2 = False):
        style = Config.cell_style
        if font2:
            cell.font = Font(**style['font2'])
        else:
            cell.font = Font(**style['font1'])
        cell.alignment = Alignment(**style['alignment'])

    def _set_rows_height(self):
        max_row = self.sheet.max_row
        for i in range(1,max_row+1):
            self.sheet.row_dimensions[i].height = Config.cell_height

    def _set_columns_width(self):
        for key, width in Config.cell_width.items():
            for col_index in DIMENSIONS[key]:
                self.sheet.column_dimensions[col_index].width = width

    def _set_cells_style(self):
        for row in self.sheet.iter_rows():
            for cell in row:
                self._set_cell_style(cell)

    def _set_title_style(self):
        title_ranges = ['B1:S1', 'T1:AK1']
        for i in range(2):
            color, background = Config.team_colors[i]
            for cell in self.sheet[title_ranges[i]][0]:
                self._set_cell_style(cell, background)
                fill = {
                    'fill_type': 'solid',
                    'fgColor': color,
                }
                cell.fill = PatternFill(**fill)

    def save(self):
        self._set_rows_height()
        self._set_columns_width()
        self._set_cells_style()
        self._set_title_style()
        self.sheet.freeze_panes = self.sheet['B2']


class Sheet:
    def __init__(self, wb, game):
        self.game = game
        self.frames = game.frames
        self.sheet = wb['sheet3']
        self.player_names = self.game.name_players
        if self.game.team_colors is None:
            Config.team_colors[0] = utils.to_hex([255, 255, 255])
            Config.team_colors[1] = utils.to_hex([70, 70, 70])
        else:
            Config.team_colors[0] = utils.to_hex(self.game.team_colors[0])
            Config.team_colors[1] = utils.to_hex(self.game.team_colors[1])


    def new(self):
        frames = self.game.frames
        self._set_title()
        for i, frame in enumerate(frames):
            self._hp_ult_charge_append(frame.players, frame.time)
        self.save()

    def _set_title(self):
        for i in range(12):
            Config.title[1 + 3 * i] = self.player_names[i]
        self.sheet.append(Config.title)

    def _hp_ult_charge_append(self, players, time):
        hp_ult_charge_row = [utils.time_format(time)]
        for player in players:
            chara = utils.chara_capitalize(player.chara)
            hp = 0 if player.is_dead else 100  # Only an indication of dead or alive now
            ult_charge = player.ult_charge
            hp_ult_charge_row += [chara, hp, ult_charge]
        self.sheet.append(hp_ult_charge_row)

    def save(self):
        """
        对 sheet 中单元格应用样式并保存
        :return: None
        """
        Save(self.sheet).save()

    def json(self):
        sheet_data = []
        sheet = self.sheet
        for col, cells in enumerate(sheet.rows):
            if col < 1:
                continue
            data = {
                'time': sheet[DIMENSIONS['time'][0] + str(col + 1)].value,
                'players': [],
            }
            for i, _ in enumerate(DIMENSIONS['players']):
                player = {
                    'chara': sheet[DIMENSIONS['players'][i] + str(col + 1)].value,
                    'hps': sheet[DIMENSIONS['hps'][i] + str(col + 1)].value,
                    'ults': sheet[DIMENSIONS['ults'][i] + str(col + 1)].value,
                    'team': self.game.team_names[0] if i < 6 else self.game.team_names[1],
                    'index': i + 1,
                }
                data['players'].append(player)
            sheet_data.append(data)
        return sheet_data
